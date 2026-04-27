#!/usr/bin/python3

import threading
import time
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class SaveData:

    def __init__(self, my_serial):
        self.serial = my_serial
        self.__dirpath = "../Data/"
        if not os.path.exists(self.__dirpath):
            os.makedirs(self.__dirpath)

    @staticmethod
    def check_path(filepath):
        if not os.path.exists(filepath):
            new_workbook = Workbook()
            new_workbook.save(filepath)

    @staticmethod
    def insert_headings(current_sheet):
        columns = ["Time", "Temperature", "Humidity", "CO2", "TVOC", "O2", "PM2.5", "PM10"]
        for i in range(8):
            cell = current_sheet.cell(column=i + 1, row=1, value=columns[i])
            cell.font = Font(size=12, bold=True)

    def save_to_local(self):
        localtime = time.localtime(time.time())
        filepath = self.__dirpath + time.strftime("%Y_%m", localtime) + ".xlsx"
        self.check_path(filepath)
        current_workbook = load_workbook(filepath)

        sheetname = time.strftime("%m-%d", localtime)
        current_sheet = current_workbook.active
        for sheet in current_workbook:
            if sheet.title == "Sheet":
                sheet.title = sheetname
                self.insert_headings(sheet)
                current_workbook.save(filepath)
                current_sheet = sheet
                break
            elif sheet.title == sheetname:
                current_sheet = sheet
                break
        if current_sheet.title != sheetname:
            current_sheet = current_workbook.create_sheet(sheetname)
            self.insert_headings(current_sheet)
            current_workbook.save(filepath)

        current_time = time.strftime("%H:%M", localtime)
        temperature = self.serial.get_temperature()
        humidity = self.serial.get_humidity()
        co2 = self.serial.get_co2()
        tvoc = self.serial.get_tvoc()
        o2 = self.serial.get_o2()
        pm25 = self.serial.get_pm25()
        pm10 = self.serial.get_pm10()
        new_data = [current_time, temperature, humidity, co2, tvoc, o2, pm25, pm10]
        current_sheet.append(new_data)
        current_workbook.save(filepath)


class AutoSave(threading.Thread):

    def __init__(self, my_serial):
        threading.Thread.__init__(self)
        self.name = "auto_save_data"
        self.serial = my_serial
        self.__save_data = SaveData(self.serial)
        self.__saved = False
        self.__closed = False

    def run(self):
        while True:
            if self.__closed:
                break
            localtime = time.localtime(time.time())
            if localtime.tm_min == 0 and not self.__saved:
                self.__save_data.save_to_local()
                self.__saved = True
            elif localtime.tm_min != 0:
                self.__saved = False

    def kill(self):
        self.__closed = True


if __name__ == '__main__':
    my_thread = AutoSave()
    my_thread.start()
    my_thread.kill()